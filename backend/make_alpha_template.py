"""
Generates Alpha_Upload_Template.xlsx — the upload template for the /alpha page.

Layout matches AlphaUploadView (backend/api/views.py):
  * Sheet named "Data"
  * Row 1 blank, Row 2 headers: Stock | Date | Price | P/E | Forward Return
  * Row 3+ daily observations
Required columns: Stock, Date, Price, P/E. Forward Return is optional
(the backend recomputes forward returns from the price series).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = Workbook()

# ---------------------------------------------------------------- Data sheet
ws = wb.active
ws.title = "Data"

HEADERS = ["Stock", "Date", "Price", "P/E", "Forward Return"]

# Row 1 left intentionally blank (backend reads header on row 2 / header=1)
HEADER_ROW = 2
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, name in enumerate(HEADERS, start=1):
    cell = ws.cell(row=HEADER_ROW, column=c, value=name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# A couple of example rows so the format is unambiguous (replace with real data)
examples = [
    ("GOOGL", date(2024, 1, 2), 139.69, 24.1, None),
    ("GOOGL", date(2024, 1, 3), 138.92, 23.9, None),
    ("GOOGL", date(2024, 1, 4), 137.39, 23.7, None),
    ("MSFT",  date(2024, 1, 2), 370.87, 34.2, None),
    ("MSFT",  date(2024, 1, 3), 370.60, 34.1, None),
]
example_font = Font(italic=True, color="808080")
for i, (stk, dt, price, pe, fwd) in enumerate(examples, start=HEADER_ROW + 1):
    ws.cell(row=i, column=1, value=stk)
    dcell = ws.cell(row=i, column=2, value=dt)
    dcell.number_format = "m/d/yyyy"
    ws.cell(row=i, column=3, value=price).number_format = "#,##0.00"
    ws.cell(row=i, column=4, value=pe).number_format = "0.0"
    # Forward Return left blank — recomputed server-side
    for c in range(1, 6):
        ws.cell(row=i, column=c).font = example_font

# Column widths + default formats for the empty rows below
widths = {"A": 12, "B": 14, "C": 12, "D": 10, "E": 16}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Pre-format the date / number columns for the next 1000 blank rows so typed
# data lands in the right format automatically.
for r in range(HEADER_ROW + 1, HEADER_ROW + 1000):
    ws.cell(row=r, column=2).number_format = "m/d/yyyy"
    ws.cell(row=r, column=3).number_format = "#,##0.00"
    ws.cell(row=r, column=4).number_format = "0.0"

ws.freeze_panes = "A3"   # keep header visible while scrolling

# ---------------------------------------------------------- Instructions sheet
info = wb.create_sheet("Instructions")
info.column_dimensions["A"].width = 100
title_font = Font(bold=True, size=14, color="1F4E79")
h_font = Font(bold=True, size=11)

lines = [
    ("Alpha Upload Template — How to use", title_font),
    ("", None),
    ("Put all your data on the 'Data' sheet, then upload it on the /alpha page "
     "(blue 'Upload Excel' button).", None),
    ("", None),
    ("Columns (header is on ROW 2 — leave row 1 blank):", h_font),
    ("   • Stock          — ticker, e.g. GOOGL  (REQUIRED)", None),
    ("   • Date           — observation date    (REQUIRED)", None),
    ("   • Price          — share price         (REQUIRED)", None),
    ("   • P/E            — price/earnings       (REQUIRED)", None),
    ("   • Forward Return — OPTIONAL. Leave blank; the dashboard recomputes it "
     "from prices for any time window you choose.", None),
    ("", None),
    ("Adding more names / data:", h_font),
    ("   • Add new tickers as more rows — just keep the same 4 columns. One row "
     "per stock per day.", None),
    ("   • Keep ONE master file with every stock and its full history.", None),
    ("", None),
    ("IMPORTANT — how upload works (full refresh per stock):", h_font),
    ("   • When you upload, every ticker that appears in the file has its old data "
     "WIPED and replaced with what's in the file.", None),
    ("   • Tickers NOT in the file are left untouched.", None),
    ("   • So: don't upload a file with only new rows for an existing stock — it "
     "would erase that stock's history. Always include the full history for any "
     "stock you upload.", None),
    ("", None),
    ("Tips:", h_font),
    ("   • The grey example rows on the Data sheet are just illustration — delete "
     "them and paste your own.", None),
    ("   • More data per stock = better statistics. Daily observations work best.", None),
    ("   • Save as .xlsx (or .xlsm if you use the UploadAlpha macro).", None),
]

r = 1
for text, font in lines:
    cell = info.cell(row=r, column=1, value=text)
    if font:
        cell.font = font
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

wb.active = 0  # open on the Data sheet
out = "Alpha_Upload_Template.xlsx"
wb.save(out)
print("wrote", out)
